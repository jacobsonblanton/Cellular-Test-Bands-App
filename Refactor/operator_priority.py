"This module will retreive the the high priority essential and non-essential band combiantion for each operator (ENDC and CA)."

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pprint import pprint
import pandas as pd

sheet = "CTIA-01.02-Operator-Priority-List-V4.0.1.xlsx"
wb = load_workbook("CTIA-01.02-Operator-Priority-List-V4.0.1.xlsx", data_only=True)
ctia_sheet = wb.active


class OperatorPriorityList(object):
    "Class for setting and retreiving the operator band priority list."

    def __init__(self) -> None:
        self.oper_ca_bands = {
            "ATT": self.att_ca_tis_trp(),
            "TMO": self.tmo_ca_tis_trp(),
            "VZW": self.vzw_ca_tis_trp(),
        }

    def get_data_tis(self) -> list:
        "A function for returning the CA TIS data."
        self.data = []
        for col in range(1, 6):
            for row in range(5, 28):
                char = get_column_letter(col)
                if ctia_sheet[char + str(row)].value != None:
                    self.data.append(ctia_sheet[char + str(row)].value)
        return self.data

    def get_data_trp(self) -> list:
        "A function for returning the CA TRP data."
        self.data = []
        for col in range(1, 6):
            for row in range(39, 68):
                char = get_column_letter(col)
                if ctia_sheet[char + str(row)].value != None:
                    self.data.append(ctia_sheet[char + str(row)].value)
        return self.data

    def att_ca_tis_trp(self) -> dict:
        "A method for retreiving the Non-Essential CA band combinations for TIS/TRP ATT."
        data = self.get_data_tis()
        att_tis_bands = {
            self.get_data_tis()[0]: {
                "TIS": [data[1:i] for i in range(len(data)) if data[i] == "Dish"][0]
            }
        }
        data = self.get_data_trp()
        att_trp_bands = {
            "TRP": [data[1:i] for i in range(len(data)) if data[i] == "Dish"][0]
        }

        return att_tis_bands[self.get_data_tis()[0]] | att_trp_bands

    def tmo_ca_tis_trp(self) -> dict:
        "A method for retreiving the Non-Essential CA band combinations for TIS/TRP TMO."
        data = self.get_data_tis()[self.get_data_tis().index("T-Mobile") + 1 :]
        tmo_tis_bands = {
            self.get_data_tis()[self.get_data_tis().index("T-Mobile") :][0]: {
                "TIS": [data[:i] for i in range(len(data)) if data[i] == "Verizon"][0]
            }
        }
        data = self.get_data_trp()[self.get_data_trp().index("T-Mobile") + 1 :]
        tmo_trp_bands = {
            "TRP": [data[:i] for i in range(len(data)) if data[i] == "Verizon"][0]
        }

        return (
            tmo_tis_bands[
                self.get_data_tis()[self.get_data_tis().index("T-Mobile") :][0]
            ]
            | tmo_trp_bands
        )

    def vzw_ca_tis_trp(self) -> dict:
        "A method for retreiving the Non-Essential CA band combinations for TIS/TRP VZW."
        vzw_tis_bands = {
            self.get_data_tis()[self.get_data_tis().index("Verizon") :][0]: {
                "TIS": self.get_data_tis()[self.get_data_tis().index("Verizon") + 1 :]
            }
        }
        vzw_trp_bands = {
            "TRP": self.get_data_trp()[self.get_data_trp().index("Verizon") + 1 :]
        }

        return (
            vzw_tis_bands[
                self.get_data_tis()[self.get_data_tis().index("Verizon") :][0]
            ]
            | vzw_trp_bands
        )


class OperatorPriorityFR1(object):
    "A class for storing the FR1 bands for each carrier."

    def __init__(self) -> None:
        self.data = self.clean_data()

    def get_data(self):
        "Returning the data from the excel sheet"
        df = pd.read_excel(
            io="CTIA-01.02-Operator-Priority-List-V4.0.1.xlsx",
            sheet_name="NonEssential High Priority",
            skiprows=range(0, 4),
            usecols=("H, J:K"),
        )
        trp = df.iloc[0:16, :]
        tis = df.iloc[34:56, :]

        return trp, tis

    def clean_data(self):
        "Formatting the data to same format as the Manufacturer Data."
        trp, tis = self.get_data()[0], self.get_data()[1]
        trp.rename(
            columns={"AT&T.1": "ATT", "T-Mobile.1": "TMO", "Verizon4": "VZW"},
            inplace=True,
        )

        tis.rename(
            columns={"AT&T.1": "ATT", "T-Mobile.1": "TMO", "Verizon4": "VZW"},
            inplace=True,
        )
        for i in range(len(trp["ATT"])):
            trp["ATT"][i] = trp["ATT"][i].replace("-", "_")
            trp["ATT"][i] = trp["ATT"][i][: trp["ATT"][i].rfind("A") + 1]

        return trp
